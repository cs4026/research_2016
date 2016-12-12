import openpyxl

wb=openpyxl.load_workbook('2016-2017.xlsx')
#sheet=wb.active
#heet.title="2016-2017 Data"
sheet=wb.active
"""sheet.cell(row=1,column=1).value='University'
sheet.cell(row=1,column=2).value='Year'
sheet.cell(row=1,column=3).value='Nation'
sheet.cell(row=1,column=4).value='UGTuition'
sheet.cell(row=1,column=5).value='Grad Tuition'"""

f=open('qrs','r')
r=2
while True:
	uni=f.readline()
	
	if not uni:
		break

	else:
		sheet.cell(row=r,column=1).value=uni
		r+=1
		print(uni)
		f.readline()		

f.close()
wb.save("2016-2017",'r')


f=open('grad new','r')
while True:
	
r=2
while True:
	
wb.save("2016-2017.xlsx")
